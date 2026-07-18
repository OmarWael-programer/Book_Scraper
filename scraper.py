import requests
import random
import time
import csv
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment , Font , PatternFill

start_time = time.perf_counter()


# Configuration
# =========================

BASE_URL = "https://books.toscrape.com/catalogue/"

CSV_FILE = r"D:\Programs\Python\projects\Book_Scraper\books.csv"

EXCEL_FILE = r"D:\Programs\Python\projects\Book_Scraper\books.xlsx"

REQUEST_TIMEOUT = 10

MAX_RETRIES = 3

HEADERS  = {"User-Agent" : ("Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36")}

FIELDNAMES = ["Title" , "Price" , "Availability" , "Rating" , "Book Link" , "UPC" , "Product Type" ,"Price (excl. tax)" , "Price (incl. tax)" , "Tax" , "Number of reviews" ,"Description"]

# ===========================


def get_page(url):
    """ download the requested page and gives 3 retries """

    session = requests.Session()
    session.headers.update(HEADERS)


    for attempt  in range(MAX_RETRIES):
        try:
            response = session.get(url , timeout= REQUEST_TIMEOUT)
            response.raise_for_status()
            response.encoding = "utf-8"
            return response       

        except Exception:
            time.sleep(random.uniform(1, 3))
            print("retry..")

    else:
        print("Failed")
        return None


def get_table_value(Book_soup , name):
    """ get the book details value """
        
    table = Book_soup.select_one("table.table.table-striped")

    details = table.find("th" , string=name)
    return details.find_next("td").text
    



def scrape_all_books(soup , page_url):
    """ get all the books data in the page """

    all_books = []

    books = soup.find_all("article" , class_="product_pod")

    for book in books:

        title = book.select_one("h3 a")["title"]


        price = book.select_one("div.product_price p.price_color").text

        
        availability = book.select_one("div.product_price p.instock.availability").text.strip()


        rating = book.select_one("p.star-rating")["class"][1]
        
        
        book_link = book.select_one("div.image_container a")["href"]
        book_link = urljoin(page_url , book_link)


        book_response = get_page(book_link)


        if book_response is None:
            print("Failed to load the book page")
            continue
        

        book_soup = BeautifulSoup(book_response.text , "html.parser")


        upc = get_table_value(book_soup , "UPC")

        product_type = get_table_value(book_soup , "Product Type")

        price_excl_tax = get_table_value(book_soup , "Price (excl. tax)")

        price_incl_tax = get_table_value(book_soup , "Price (incl. tax)")

        tax = get_table_value(book_soup , "Tax")

        number_of_reviews = get_table_value(book_soup , "Number of reviews")

        description = book_soup.select_one("div.sub-header").find_next("p").text
        description = description.replace("\u2028" , "").replace("\u2029" , "").replace("\n" , "").replace("\r" ,"")
        

        all_books.append({
            "Title" : title,
            "Price" : price,
            "Availability" : availability,
            "Rating" : rating,
            "Book Link" : book_link,
            "UPC" : upc,
            "Product Type" : product_type,
            "Price (excl. tax)" : price_excl_tax,
            "Price (incl. tax)" : price_incl_tax,
            "Tax" : tax,
            "Number of reviews" : number_of_reviews,
            "Description" : description
        })

    return all_books
       
    
def save_csv(data):
    """ saves the data to a csv file with the header """

    with open(CSV_FILE , "w" , newline="" , encoding="utf-8") as file:

        writer = csv.DictWriter(
            file,
            fieldnames= FIELDNAMES
        )

        writer.writeheader()


        writer.writerows(data)

def save_excel(data):
    """ saves the data to a excel file with the header format """
    
    workbook = Workbook()

    sheet = workbook.active

    sheet.append(FIELDNAMES)

    for book in data:

        for fieldname in FIELDNAMES:
            sheet.append([book[fieldname]])



    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 50
    sheet.column_dimensions["F"].width = 35
    sheet.column_dimensions["G"].width = 10
    sheet.column_dimensions["H"].width = 15
    sheet.column_dimensions["I"].width = 15
    sheet.column_dimensions["J"].width = 10
    sheet.column_dimensions["K"].width = 10
    sheet.column_dimensions["L"].width = 100


    header_row = 1


    header_font = Font(bold=True)

    for cell in sheet[header_row]:
        cell.font = header_font
        


    header_fill = PatternFill(
        fill_type="solid",
        start_color="D9D9D9"
    )

    for cell in sheet[header_row]:
        cell.fill = header_fill


    
    alignment = Alignment(horizontal="center")

    for cell in sheet[header_row]:
        cell.alignment = alignment


    freeze_row = "A2"
    sheet.freeze_panes = freeze_row


    maximum_row = sheet.max_row

    sheet.auto_filter.ref = f"A1:L{maximum_row}"


    workbook.save(EXCEL_FILE)





books_found = 0
page_number = 1

url = urljoin(BASE_URL, f"page-{page_number}.html")

all_books = []

while True:

    book_collected = 0
    
    response = get_page(url)

    if response is None:
        print("Failed to load:")
        print(url)
        page_number += 1
        continue
    
    soup = BeautifulSoup(response.text , "html.parser")

    
    page_books = scrape_all_books(soup , url)
    all_books.extend(page_books)

    for book in page_books:
        book_collected += 1
        books_found += 1
    
    print(f"Page {page_number}...")
    print(f"{book_collected} Books collected")



    next_button = soup.select_one("li.next")

    if next_button is None:
        break
    
    url = urljoin(url , next_button.select_one("a")["href"])
    page_number += 1

save_csv(all_books)
save_excel(all_books)

end_time = time.perf_counter()

print("===== Statistics =====")
print(f"Pages scanned : {page_number}")
print(f"Books found   : {books_found}")
print(f"Elapsed time  : {(end_time - start_time) / 60:.3} minutes")

    